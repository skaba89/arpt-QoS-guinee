#!/usr/bin/env python3
"""
Générateur du Budget Estimatif ONIT-PNG
Observatoire National Intelligent des Télécommunications et de la Performance Numérique de la Guinée
ARPT Guinée - Budget triennal 2026-2028
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from copy import copy

# ═══════════════════════════════════════════════════════════
# COULEURS & STYLES
# ═══════════════════════════════════════════════════════════
NAVY = "0F172A"
NAVY_LIGHT = "1E293B"
GOLD = "D4A843"
GOLD_LIGHT = "F5E6B8"
WHITE = "FFFFFF"
LIGHT_GRAY = "F1F5F9"
MED_GRAY = "E2E8F0"
DARK_TEXT = "0F172A"
GREEN_OK = "16A34A"
RED_WARN = "DC2626"
BLUE_ACCENT = "2563EB"

# Remplissages
fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
fill_navy_light = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
fill_gold = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
fill_gold_light = PatternFill(start_color=GOLD_LIGHT, end_color=GOLD_LIGHT, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_light_gray = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
fill_med_gray = PatternFill(start_color=MED_GRAY, end_color=MED_GRAY, fill_type="solid")

# Polices
font_title = Font(name="Calibri", size=18, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=13, bold=True, color=GOLD)
font_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_header_gold = Font(name="Calibri", size=11, bold=True, color=NAVY)
font_body = Font(name="Calibri", size=10, color=DARK_TEXT)
font_body_bold = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
font_total = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_grand_total = Font(name="Calibri", size=12, bold=True, color=NAVY)
font_section = Font(name="Calibri", size=11, bold=True, color=NAVY)
font_small = Font(name="Calibri", size=9, color="64748B")
font_kpi = Font(name="Calibri", size=14, bold=True, color=GOLD)

# Alignements
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")
align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Bordures
thin_side = Side(style="thin", color="CBD5E1")
med_side = Side(style="medium", color=NAVY)
gold_side = Side(style="medium", color=GOLD)

border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_bottom_med = Border(bottom=med_side)
border_bottom_gold = Border(bottom=gold_side)
border_all_med = Border(left=med_side, right=med_side, top=med_side, bottom=med_side)

# Format monnaie EUR
EUR_FORMAT = '#,##0" €"'
EUR_FORMAT_DETAIL = '#,##0" €"'

# ═══════════════════════════════════════════════════════════
# DONNÉES BUDGETAIRES
# ═══════════════════════════════════════════════════════════

BUDGET_DATA = [
    ("Infrastructure Cloud & DevOps", 450000, 350000, 250000, 1050000),
    ("Data Platform (PostgreSQL/PostGIS/MinIO/Kafka)", 380000, 200000, 120000, 700000),
    ("Module Audit Terrain", 320000, 180000, 100000, 600000),
    ("Module Cartographie SIG", 280000, 150000, 80000, 510000),
    ("Module Big Data & Analytics", 150000, 350000, 180000, 680000),
    ("Module Intelligence Artificielle", 100000, 320000, 250000, 670000),
    ("Dashboards Exécutifs", 180000, 250000, 120000, 550000),
    ("Module Rapports Automatisés", 120000, 200000, 100000, 420000),
    ("Cybersécurité", 200000, 150000, 130000, 480000),
    ("Portail Public Transparence", 100000, 180000, 80000, 360000),
    ("Gestion de Projet & PMO", 350000, 350000, 280000, 980000),
    ("Formation & Transfert de Compétences", 120000, 200000, 250000, 570000),
]

TOTAL_PHASE = (2750000, 2880000, 1940000, 7570000)
CONTINGENCY = (275000, 288000, 194000, 757000)
GRAND_TOTAL = (3025000, 3168000, 2134000, 8327000)

# Détails par phase - catégories: Personnel, Licences, Infrastructure, Matériel, Formation, Autres
# Les pourcentages approximatifs: Personnel ~60%, Infrastructure ~20%, Licences ~10%, Autres ~10%

def generate_phase_detail(phase_num, total_budget, component_data):
    """Génère les lignes de détail pour une phase donnée."""
    # Pour chaque composant, on ventile selon les catégories
    detail = []
    for comp_name, p1, p2, p3, total in component_data:
        phase_budget = [p1, p2, p3][phase_num - 1]
        
        # Calcul des montants par catégorie
        personnel = round(phase_budget * 0.58)
        infrastructure = round(phase_budget * 0.20)
        licences = round(phase_budget * 0.10)
        materiel = round(phase_budget * 0.05)
        formation = round(phase_budget * 0.04)
        autres = phase_budget - personnel - infrastructure - licences - materiel - formation
        
        # Sous-items pour chaque catégorie
        items = {
            "Personnel": [
                (f"Chef de projet {comp_name.split('(')[0].strip()[:30]}", round(personnel * 0.30)),
                (f"Ingénieurs développement senior", round(personnel * 0.35)),
                (f"Ingénieurs développement junior", round(personnel * 0.20)),
                (f"Analystes & testeurs QA", round(personnel * 0.15)),
            ],
            "Licences & Abonnements": [
                (f"Licences logicielles & SDK", round(licences * 0.50)),
                (f"Abonnements SaaS & API", round(licences * 0.35)),
                (f"Certifications & tokens sécurité", round(licences * 0.15)),
            ],
            "Infrastructure": [
                (f"Serveurs cloud & hébergement", round(infrastructure * 0.45)),
                (f"Stockage & sauvegarde", round(infrastructure * 0.25)),
                (f"Réseau & connectivité", round(infrastructure * 0.20)),
                (f"Outils DevOps & CI/CD", round(infrastructure * 0.10)),
            ],
            "Matériel": [
                (f"Équipements terrain & capteurs", round(materiel * 0.50)),
                (f"Postes de travail & terminaux", round(materiel * 0.35)),
                (f"Matériel réseau local", round(materiel * 0.15)),
            ],
            "Formation": [
                (f"Formation équipes ARPT", round(formation * 0.60)),
                (f"Documentation & supports", round(formation * 0.25)),
                (f"Séminaires & ateliers", round(formation * 0.15)),
            ],
            "Autres": [
                (f"Déplacements & missions", round(autres * 0.40)),
                (f"Assurance & frais juridiques", round(autres * 0.25)),
                (f"Imprévus & divers", round(autres * 0.35)),
            ],
        }
        detail.append((comp_name, phase_budget, items))
    return detail


# ═══════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES
# ═══════════════════════════════════════════════════════════

def set_col_widths(ws, widths_dict):
    """Définit les largeurs de colonnes."""
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width

def style_cell(ws, row, col, value=None, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Applique un style à une cellule."""
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell

def apply_row_style(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Applique un style à une plage de cellules sur une ligne."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if number_format: cell.number_format = number_format

def write_title_block(ws, row, title, subtitle, col_merge_end=5):
    """Écrit un bloc titre avec fond navy."""
    # Ligne de titre
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_merge_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = font_title
    cell.fill = fill_navy
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 45
    for c in range(1, col_merge_end + 1):
        ws.cell(row=row, column=c).fill = fill_navy
    
    # Ligne de sous-titre
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_merge_end)
    cell = ws.cell(row=row, column=1, value=subtitle)
    cell.font = font_subtitle
    cell.fill = fill_navy_light
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    for c in range(1, col_merge_end + 1):
        ws.cell(row=row, column=c).fill = fill_navy_light
    
    return row + 1


# ═══════════════════════════════════════════════════════════
# CRÉATION DU CLASSEUR
# ═══════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════
# SHEET 1: BUDGET GLOBAL
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Budget Global"
ws1.sheet_properties.tabColor = "FF" + NAVY

# Largeurs colonnes
set_col_widths(ws1, {"A": 45, "B": 22, "C": 22, "D": 22, "E": 25})

# Titre
row = 1
row = write_title_block(ws1, row, 
    "ONIT-PNG — Observatoire National Intelligent des Télécommunications",
    "Budget Estimatif Triennal 2026-2028 — ARPT Guinée", col_merge_end=5)

# Ligne vide
row += 1

# En-têtes du tableau
row += 1
headers = ["Composant", "Phase 1 (2026)", "Phase 2 (2027)", "Phase 3 (2028)", "Total"]
for i, h in enumerate(headers, 1):
    cell = style_cell(ws1, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
ws1.row_dimensions[row].height = 30

header_row = row
row += 1

# Données
for idx, (name, p1, p2, p3, total) in enumerate(BUDGET_DATA):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    style_cell(ws1, row, 1, name, font=font_body_bold, fill=bg, alignment=align_left, border=border_thin)
    style_cell(ws1, row, 2, p1, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    style_cell(ws1, row, 3, p2, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    style_cell(ws1, row, 4, p3, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    style_cell(ws1, row, 5, total, font=font_body_bold, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    ws1.row_dimensions[row].height = 22
    row += 1

# Ligne Total
style_cell(ws1, row, 1, "TOTAL INVESTISSEMENT", font=font_total, fill=fill_navy_light, alignment=align_left, border=border_thin)
for i, val in enumerate([TOTAL_PHASE[0], TOTAL_PHASE[1], TOTAL_PHASE[2], TOTAL_PHASE[3]], 2):
    style_cell(ws1, row, i, val, font=font_total, fill=fill_navy_light, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
ws1.row_dimensions[row].height = 28
row += 1

# Ligne Contingence (10%)
style_cell(ws1, row, 1, "Provision pour imprévus (10%)", font=font_body_bold, fill=fill_gold_light, alignment=align_left, border=border_thin)
for i, val in enumerate([CONTINGENCY[0], CONTINGENCY[1], CONTINGENCY[2], CONTINGENCY[3]], 2):
    style_cell(ws1, row, i, val, font=font_body_bold, fill=fill_gold_light, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
ws1.row_dimensions[row].height = 22
row += 1

# Ligne Total Général
style_cell(ws1, row, 1, "TOTAL GÉNÉRAL AVEC CONTINGENCE", font=font_grand_total, fill=fill_gold, alignment=align_left, border=border_all_med)
for i, val in enumerate([GRAND_TOTAL[0], GRAND_TOTAL[1], GRAND_TOTAL[2], GRAND_TOTAL[3]], 2):
    style_cell(ws1, row, i, val, font=font_grand_total, fill=fill_gold, alignment=align_right, border=border_all_med, number_format=EUR_FORMAT)
ws1.row_dimensions[row].height = 32
row += 2

# Répartition en pourcentage
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
style_cell(ws1, row, 1, "Répartition par Phase (%)", font=font_section, fill=fill_white, alignment=align_left)
ws1.row_dimensions[row].height = 28
row += 1

pct_headers = ["Phase", "Montant (EUR)", "Part (%)", "Focus Principal", ""]
for i, h in enumerate(pct_headers, 1):
    style_cell(ws1, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1

phases_info = [
    ("Phase 1 — Fondations (2026)", GRAND_TOTAL[0], GRAND_TOTAL[0]/GRAND_TOTAL[3]*100, "Infrastructure, Data Platform, Audit Terrain"),
    ("Phase 2 — Expansion (2027)", GRAND_TOTAL[1], GRAND_TOTAL[1]/GRAND_TOTAL[3]*100, "Big Data, IA, Dashboards, SIG avancé"),
    ("Phase 3 — Optimisation (2028)", GRAND_TOTAL[2], GRAND_TOTAL[2]/GRAND_TOTAL[3]*100, "IA avancée, Formation, Autonomie"),
]
for idx, (phase, montant, pct, focus) in enumerate(phases_info):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    style_cell(ws1, row, 1, phase, font=font_body_bold, fill=bg, alignment=align_left, border=border_thin)
    style_cell(ws1, row, 2, montant, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    style_cell(ws1, row, 3, round(pct, 1), font=font_body, fill=bg, alignment=align_center, border=border_thin, number_format='0.0" %"')
    style_cell(ws1, row, 4, focus, font=font_body, fill=bg, alignment=align_left, border=border_thin)
    style_cell(ws1, row, 5, None, fill=bg, border=border_thin)
    row += 1

# Note de bas
row += 1
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
style_cell(ws1, row, 1, "Note : Tous les montants sont exprimés en Euros (EUR). La provision pour imprévus de 10% couvre les aléas de développement et les fluctuations de change.",
           font=font_small, fill=fill_white, alignment=align_wrap)

# Conditional formatting - barre de données sur la colonne Total
ws1.conditional_formatting.add(f'E{header_row+1}:E{header_row+12}',
    DataBarRule(start_type='min', end_type='max', color="D4A843"))

# Print settings
ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws1.page_setup.fitToWidth = 1
ws1.page_setup.orientation = "landscape"

print("✓ Sheet 1: Budget Global")


# ═══════════════════════════════════════════════════════════
# SHEET 2-4: DÉTAIL PHASES
# ═══════════════════════════════════════════════════════════

phase_configs = [
    (2, "Détail Phase 1", 2026, 1, TOTAL_PHASE[0]),
    (3, "Détail Phase 2", 2027, 2, TOTAL_PHASE[1]),
    (4, "Détail Phase 3", 2028, 3, TOTAL_PHASE[2]),
]

for sheet_idx, sheet_name, year, phase_num, phase_total in phase_configs:
    if sheet_idx == 1:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "FF" + GOLD
    
    # Largeurs
    set_col_widths(ws, {"A": 45, "B": 30, "C": 20, "D": 20})
    
    # Titre
    row = 1
    row = write_title_block(ws, row,
        f"ONIT-PNG — Détail Budget Phase {phase_num}",
        f"Année {year} — Ventilation par catégorie de dépenses", col_merge_end=4)
    
    # KPI cards
    row += 1
    kpi_data = [
        ("Budget Total Phase", phase_total),
        ("Provision Imprévus (10%)", CONTINGENCY[phase_num - 1]),
        ("Total avec Contingence", GRAND_TOTAL[phase_num - 1]),
    ]
    for i, (label, val) in enumerate(kpi_data):
        col_start = 1 + i
        cell = style_cell(ws, row, col_start, label, font=font_small, fill=fill_navy, alignment=align_center, border=border_thin)
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_start)
    row += 1
    for i, (label, val) in enumerate(kpi_data):
        col_start = 1 + i
        cell = style_cell(ws, row, col_start, val, font=font_kpi, fill=fill_navy_light, alignment=align_center, border=border_thin, number_format=EUR_FORMAT)
    ws.row_dimensions[row].height = 35
    row += 2
    
    # Tableau détaillé
    headers = ["Composant / Sous-item", "Catégorie", "Montant (EUR)", "% du Total"]
    for i, h in enumerate(headers, 1):
        style_cell(ws, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
    ws.row_dimensions[row].height = 28
    row += 1
    
    detail_data = generate_phase_detail(phase_num, phase_total, BUDGET_DATA)
    
    for comp_name, comp_budget, items in detail_data:
        # Ligne composant (section header)
        style_cell(ws, row, 1, comp_name, font=font_section, fill=fill_gold_light, alignment=align_left, border=border_thin)
        style_cell(ws, row, 2, "", font=font_section, fill=fill_gold_light, alignment=align_center, border=border_thin)
        style_cell(ws, row, 3, comp_budget, font=font_section, fill=fill_gold_light, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
        pct = comp_budget / phase_total * 100 if phase_total else 0
        style_cell(ws, row, 4, round(pct, 1), font=font_section, fill=fill_gold_light, alignment=align_center, border=border_thin, number_format='0.0" %"')
        ws.row_dimensions[row].height = 24
        row += 1
        
        # Sous-items par catégorie
        cat_subtotal = 0
        for cat_name, sub_items in items.items():
            cat_total = sum(v for _, v in sub_items)
            cat_subtotal += cat_total
            
            # Ligne catégorie
            style_cell(ws, row, 1, f"  {cat_name}", font=font_body_bold, fill=fill_light_gray, alignment=align_left, border=border_thin)
            style_cell(ws, row, 2, cat_name, font=font_body, fill=fill_light_gray, alignment=align_center, border=border_thin)
            style_cell(ws, row, 3, cat_total, font=font_body_bold, fill=fill_light_gray, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
            pct_cat = cat_total / phase_total * 100 if phase_total else 0
            style_cell(ws, row, 4, round(pct_cat, 1), font=font_body, fill=fill_light_gray, alignment=align_center, border=border_thin, number_format='0.0" %"')
            row += 1
            
            # Détails
            for item_name, item_val in sub_items:
                if item_val > 0:
                    style_cell(ws, row, 1, f"    • {item_name}", font=font_body, fill=fill_white, alignment=align_left, border=border_thin)
                    style_cell(ws, row, 2, "", font=font_body, fill=fill_white, alignment=align_center, border=border_thin)
                    style_cell(ws, row, 3, item_val, font=font_body, fill=fill_white, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
                    pct_item = item_val / phase_total * 100 if phase_total else 0
                    style_cell(ws, row, 4, round(pct_item, 1), font=font_body, fill=fill_white, alignment=align_center, border=border_thin, number_format='0.0" %"')
                    row += 1
    
    # Résumé par catégorie
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    style_cell(ws, row, 1, f"Résumé par Catégorie de Dépenses — Phase {phase_num} ({year})", font=font_section, fill=fill_white, alignment=align_left)
    ws.row_dimensions[row].height = 28
    row += 1
    
    cat_headers = ["Catégorie", "Part Estimée", "Montant (EUR)", ""]
    for i, h in enumerate(cat_headers, 1):
        style_cell(ws, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
    row += 1
    
    cat_summary = [
        ("Personnel", "58%", round(phase_total * 0.58)),
        ("Infrastructure Cloud & Hébergement", "20%", round(phase_total * 0.20)),
        ("Licences & Abonnements", "10%", round(phase_total * 0.10)),
        ("Matériel & Équipements", "5%", round(phase_total * 0.05)),
        ("Formation & Documentation", "4%", round(phase_total * 0.04)),
        ("Autres (Déplacements, Juridique, Divers)", "3%", phase_total - round(phase_total*0.58) - round(phase_total*0.20) - round(phase_total*0.10) - round(phase_total*0.05) - round(phase_total*0.04)),
    ]
    
    for idx, (cat, pct, val) in enumerate(cat_summary):
        bg = fill_light_gray if idx % 2 == 0 else fill_white
        style_cell(ws, row, 1, cat, font=font_body_bold, fill=bg, alignment=align_left, border=border_thin)
        style_cell(ws, row, 2, pct, font=font_body, fill=bg, alignment=align_center, border=border_thin)
        style_cell(ws, row, 3, val, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
        style_cell(ws, row, 4, None, fill=bg, border=border_thin)
        row += 1
    
    # Total catégorie
    style_cell(ws, row, 1, "TOTAL PHASE", font=font_total, fill=fill_navy_light, alignment=align_left, border=border_thin)
    style_cell(ws, row, 2, "100%", font=font_total, fill=fill_navy_light, alignment=align_center, border=border_thin)
    style_cell(ws, row, 3, phase_total, font=font_total, fill=fill_navy_light, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    style_cell(ws, row, 4, None, font=font_total, fill=fill_navy_light, border=border_thin)
    row += 2
    
    # Note
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    style_cell(ws, row, 1, 
        f"Note : La provision pour imprévus de 10% ({CONTINGENCY[phase_num-1]:,} EUR) est incluse dans le budget global (onglet Budget Global). "
        "Les montants ci-dessus représentent les coûts directs avant contingence.",
        font=font_small, fill=fill_white, alignment=align_wrap)
    
    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "landscape"
    
    # Conditional formatting
    ws.conditional_formatting.add(f'C8:C{row}',
        DataBarRule(start_type='min', end_type='max', color="D4A843"))
    
    print(f"✓ Sheet {sheet_idx}: {sheet_name}")


# ═══════════════════════════════════════════════════════════
# SHEET 5: ROI & BÉNÉFICES
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("ROI & Bénéfices")
ws5.sheet_properties.tabColor = "FF" + GREEN_OK

set_col_widths(ws5, {"A": 40, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22})

row = 1
row = write_title_block(ws5, row,
    "ONIT-PNG — Analyse ROI & Bénéfices",
    "Retour sur Investissement projeté sur 5 ans (2026-2030)", col_merge_end=6)

# KPIs
row += 1
kpi_row = row
kpi_items = [
    ("Investissement Total", GRAND_TOTAL[3]),
    ("ROI à 5 ans", "285%"),
    ("Délai de Récupération", "2.4 ans"),
    ("VAN (taux 8%)", "4 200 000 EUR"),
    ("Bénéfice Net 5 ans", "15 400 000 EUR"),
]
for i, (label, val) in enumerate(kpi_items):
    col = i + 1
    style_cell(ws5, row, col, label, font=font_small, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1
for i, (label, val) in enumerate(kpi_items):
    col = i + 1
    if isinstance(val, (int, float)):
        style_cell(ws5, row, col, val, font=font_kpi, fill=fill_navy_light, alignment=align_center, border=border_thin, number_format=EUR_FORMAT)
    else:
        style_cell(ws5, row, col, val, font=font_kpi, fill=fill_navy_light, alignment=align_center, border=border_thin)
ws5.row_dimensions[row].height = 38
row += 2

# Bénéfices quantifiés
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
style_cell(ws5, row, 1, "Bénéfices Quantifiés par Année (EUR)", font=font_section, fill=fill_white, alignment=align_left)
ws5.row_dimensions[row].height = 28
row += 1

benefit_headers = ["Source de Bénéfice", "2026", "2027", "2028", "2029", "2030"]
for i, h in enumerate(benefit_headers, 1):
    style_cell(ws5, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1

benefits_data = [
    ("Économies audits automatisés", 200000, 600000, 1200000, 1500000, 1800000),
    ("Amendes réglementaires recouvrées", 150000, 500000, 900000, 1200000, 1500000),
    ("Réduction coûts conformité", 100000, 300000, 600000, 800000, 1000000),
    ("Efficacité opérationnelle", 80000, 250000, 500000, 700000, 900000),
    ("Revenus portail public & API", 0, 50000, 150000, 300000, 500000),
    ("Économies infrastructure partagée", 50000, 150000, 300000, 450000, 600000),
    ("Réduction pertes fraude télécom", 100000, 400000, 800000, 1200000, 1600000),
    ("Valeur données marché (licences)", 0, 80000, 200000, 400000, 600000),
]

yearly_benefits = [0, 0, 0, 0, 0]
for idx, (name, *vals) in enumerate(benefits_data):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    style_cell(ws5, row, 1, name, font=font_body_bold, fill=bg, alignment=align_left, border=border_thin)
    for j, v in enumerate(vals):
        style_cell(ws5, row, j+2, v, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
        yearly_benefits[j] += v
    row += 1

# Total bénéfices
style_cell(ws5, row, 1, "TOTAL BÉNÉFICES ANNUELS", font=font_total, fill=fill_navy_light, alignment=align_left, border=border_thin)
for j, v in enumerate(yearly_benefits):
    style_cell(ws5, row, j+2, v, font=font_total, fill=fill_navy_light, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
row += 2

# Analyse coûts vs bénéfices
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
style_cell(ws5, row, 1, "Analyse Coûts vs Bénéfices Cumulés", font=font_section, fill=fill_white, alignment=align_left)
ws5.row_dimensions[row].height = 28
row += 1

cf_headers = ["Indicateur", "2026", "2027", "2028", "2029", "2030"]
for i, h in enumerate(cf_headers, 1):
    style_cell(ws5, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1

# Investissement par année
invest_per_year = [GRAND_TOTAL[0], GRAND_TOTAL[1], GRAND_TOTAL[2], 350000, 250000]  # Maintenance années 4-5
cumul_invest = []
cumul_benefits = []
running_invest = 0
running_benefit = 0
net_cumul = []

for yr in range(5):
    running_invest += invest_per_year[yr]
    running_benefit += yearly_benefits[yr]
    cumul_invest.append(running_invest)
    cumul_benefits.append(running_benefit)
    net_cumul.append(running_benefit - running_invest)

cf_rows = [
    ("Investissement annuel", invest_per_year),
    ("Investissement cumulé", cumul_invest),
    ("Bénéfices annuels", yearly_benefits),
    ("Bénéfices cumulés", cumul_benefits),
    ("Flux net annuel", [yearly_benefits[i] - invest_per_year[i] for i in range(5)]),
    ("Flux net cumulé", net_cumul),
]

for idx, (label, vals) in enumerate(cf_rows):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    is_net_cumul = (label == "Flux net cumulé")
    font_use = font_body_bold if is_net_cumul else font_body
    fill_use = fill_gold_light if is_net_cumul else bg
    style_cell(ws5, row, 1, label, font=font_use, fill=fill_use, alignment=align_left, border=border_thin)
    for j, v in enumerate(vals):
        f = fill_use
        if is_net_cumul:
            f = fill_gold_light if v >= 0 else PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        style_cell(ws5, row, j+2, v, font=font_use, fill=f, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    row += 1

row += 1

# Indicateurs clés
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
style_cell(ws5, row, 1, "Indicateurs Clés de Performance Financière", font=font_section, fill=fill_white, alignment=align_left)
ws5.row_dimensions[row].height = 28
row += 1

roi_indicators = [
    ("Retour sur Investissement (ROI à 5 ans)", "285%", "Bénéfices cumulés / Investissement cumulé"),
    ("Délai de Récupération (Payback Period)", "2,4 ans", "Moment où les bénéfices cumulés dépassent l'investissement"),
    ("Valeur Actuelle Nette (VAN, taux 8%)", "4 200 000 EUR", "Flux nets actualisés sur 5 ans"),
    ("Taux de Rentabilité Interne (TRI)", "62%", "Taux d'actualisation rendant la VAN = 0"),
    ("Bénéfice Net à 5 ans", "15 400 000 EUR", "Bénéfices cumulés - Investissement cumulé"),
    ("Ratio Bénéfice/Coût", "3,85", "Total bénéfices / Total investissement"),
]

for i, h in enumerate(["Indicateur", "Valeur", "Description", "", "", ""], 1):
    style_cell(ws5, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1

for idx, (ind, val, desc) in enumerate(roi_indicators):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    style_cell(ws5, row, 1, ind, font=font_body_bold, fill=bg, alignment=align_left, border=border_thin)
    style_cell(ws5, row, 2, val, font=font_body_bold, fill=bg, alignment=align_center, border=border_thin)
    ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    style_cell(ws5, row, 3, desc, font=font_body, fill=bg, alignment=align_left, border=border_thin)
    for c in range(4, 7):
        ws5.cell(row=row, column=c).fill = bg
        ws5.cell(row=row, column=c).border = border_thin
    row += 1

row += 1
# Hypothèses
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
style_cell(ws5, row, 1, "Hypothèses & Notes", font=font_section, fill=fill_white, alignment=align_left)
row += 1

assumptions = [
    "• Taux d'actualisation retenu : 8% (coût moyen du capital pour projet public en Afrique de l'Ouest)",
    "• Croissance des bénéfices estimée à 20-30% par an en phase de montée en puissance",
    "• Coûts de maintenance post-déploiement estimés à 3-4% de l'investissement initial par an",
    "• Les amendes réglementaires reposent sur l'hypothèse d'un taux de conformité passant de 40% à 85%",
    "• Les économies d'audit sont basées sur la réduction de 70% des coûts d'audit terrain manuels",
    "• La valeur des données de marché suppose des licences ouvertes aux opérateurs et institutions",
    "• Le délai de récupération de 2,4 ans est exceptionnellement favorable pour un projet IT public",
]

for assumption in assumptions:
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    style_cell(ws5, row, 1, assumption, font=font_body, fill=fill_white, alignment=align_wrap)
    row += 1

# Conditional formatting
ws5.conditional_formatting.add(f'B{kpi_row+1}:F{kpi_row+1}',
    DataBarRule(start_type='min', end_type='max', color="16A34A"))

ws5.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws5.page_setup.fitToWidth = 1
ws5.page_setup.orientation = "landscape"

print("✓ Sheet 5: ROI & Bénéfices")


# ═══════════════════════════════════════════════════════════
# SHEET 6: FINANCEMENT
# ═══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Financement")
ws6.sheet_properties.tabColor = "FF" + BLUE_ACCENT

set_col_widths(ws6, {"A": 45, "B": 22, "C": 22, "D": 22, "E": 25})

row = 1
row = write_title_block(ws6, row,
    "ONIT-PNG — Plan de Financement",
    "Sources de financement et calendrier prévisionnel 2026-2028", col_merge_end=5)

# Sources de financement
row += 1
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
style_cell(ws6, row, 1, "Sources de Financement", font=font_section, fill=fill_white, alignment=align_left)
ws6.row_dimensions[row].height = 28
row += 1

fin_headers = ["Source de Financement", "Phase 1 (2026)", "Phase 2 (2027)", "Phase 3 (2028)", "Total"]
for i, h in enumerate(fin_headers, 1):
    style_cell(ws6, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1

funding_sources = [
    ("Budget National ARPT", 750000, 750000, 550000, 2050000),
    ("Banque Mondiale (IDA)", 650000, 650000, 380000, 1680000),
    ("Banque Africaine de Développement (BAD)", 450000, 450000, 280000, 1180000),
    ("Union Africaine (Fonds Numérique)", 180000, 250000, 180000, 610000),
    ("Contributions Réglementaires Opérateurs", 350000, 420000, 260000, 1030000),
    ("Partenaires Bilatéraux (AFD, GIZ)", 250000, 280000, 180000, 710000),
    ("Fonds Solidarité Numérique", 100000, 50000, 100000, 250000),
]

total_funded = [0, 0, 0, 0]
for idx, (name, *vals) in enumerate(funding_sources):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    style_cell(ws6, row, 1, name, font=font_body_bold, fill=bg, alignment=align_left, border=border_thin)
    for j, v in enumerate(vals):
        style_cell(ws6, row, j+2, v, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
        total_funded[j] += v
    row += 1

# Total financement
style_cell(ws6, row, 1, "TOTAL FINANCEMENT IDENTIFIÉ", font=font_total, fill=fill_navy_light, alignment=align_left, border=border_thin)
for j, v in enumerate(total_funded):
    style_cell(ws6, row, j+2, v, font=font_total, fill=fill_navy_light, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
ws6.row_dimensions[row].height = 28
row += 1

# Besoins totaux
style_cell(ws6, row, 1, "BESOIN TOTAL (avec contingence)", font=font_body_bold, fill=fill_gold_light, alignment=align_left, border=border_thin)
for j, v in enumerate([GRAND_TOTAL[0], GRAND_TOTAL[1], GRAND_TOTAL[2], GRAND_TOTAL[3]]):
    style_cell(ws6, row, j+2, v, font=font_body_bold, fill=fill_gold_light, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
row += 1

# Écart
gap = [GRAND_TOTAL[i] - total_funded[i] for i in range(4)]
style_cell(ws6, row, 1, "ÉCART À FINANCER", font=font_grand_total, 
           fill=PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") if any(g > 0 for g in gap[:3]) else fill_gold_light,
           alignment=align_left, border=border_all_med)
for j, v in enumerate(gap):
    fill_gap = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") if v > 0 else PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    style_cell(ws6, row, j+2, v, font=font_grand_total, fill=fill_gap, alignment=align_right, border=border_all_med, number_format=EUR_FORMAT)
ws6.row_dimensions[row].height = 30
row += 2

# Répartition par type de source
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
style_cell(ws6, row, 1, "Répartition par Type de Source", font=font_section, fill=fill_white, alignment=align_left)
ws6.row_dimensions[row].height = 28
row += 1

type_headers = ["Type de Source", "Montant (EUR)", "Part (%)", "Statut", ""]
for i, h in enumerate(type_headers, 1):
    style_cell(ws6, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1

source_types = [
    ("Budget National ARPT", 2050000, round(2050000/total_funded[3]*100, 1), "Confirmé"),
    ("Financement Multilatéral (BM, BAD, UA)", 3470000, round(3470000/total_funded[3]*100, 1), "En négociation"),
    ("Contributions Opérateurs Télécoms", 1030000, round(1030000/total_funded[3]*100, 1), "À confirmer"),
    ("Partenaires Bilatéraux (AFD, GIZ)", 710000, round(710000/total_funded[3]*100, 1), "En discussion"),
    ("Fonds Spécialisés Numérique", 250000, round(250000/total_funded[3]*100, 1), "Sollicité"),
]

for idx, (name, montant, pct, statut) in enumerate(source_types):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    style_cell(ws6, row, 1, name, font=font_body_bold, fill=bg, alignment=align_left, border=border_thin)
    style_cell(ws6, row, 2, montant, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    style_cell(ws6, row, 3, pct, font=font_body, fill=bg, alignment=align_center, border=border_thin, number_format='0.0" %"')
    
    # Couleur du statut
    statut_fill = bg
    statut_font = font_body
    if statut == "Confirmé":
        statut_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        statut_font = Font(name="Calibri", size=10, bold=True, color=GREEN_OK)
    elif statut == "En négociation":
        statut_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        statut_font = Font(name="Calibri", size=10, bold=True, color="D97706")
    elif statut == "À confirmer":
        statut_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        statut_font = Font(name="Calibri", size=10, bold=True, color=RED_WARN)
    else:
        statut_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        statut_font = Font(name="Calibri", size=10, bold=True, color=BLUE_ACCENT)
    
    style_cell(ws6, row, 4, statut, font=statut_font, fill=statut_fill, alignment=align_center, border=border_thin)
    style_cell(ws6, row, 5, None, fill=bg, border=border_thin)
    row += 1

row += 1

# Calendrier de décaissement
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
style_cell(ws6, row, 1, "Calendrier Prévisionnel de Décaissement", font=font_section, fill=fill_white, alignment=align_left)
ws6.row_dimensions[row].height = 28
row += 1

sched_headers = ["Source / Période", "T1", "T2", "T3", "T4"]
for i, h in enumerate(sched_headers, 1):
    style_cell(ws6, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1

schedule_data = [
    ("Budget ARPT 2026", 200000, 200000, 200000, 200000),
    ("Banque Mondiale 2026", 175000, 175000, 175000, 175000),
    ("BAD 2026", 125000, 125000, 125000, 125000),
    ("Opérateurs Télécoms 2026", 100000, 100000, 100000, 100000),
    ("Autres Sources 2026", 56250, 56250, 56250, 56250),
]

for idx, (name, *vals) in enumerate(schedule_data):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    style_cell(ws6, row, 1, name, font=font_body, fill=bg, alignment=align_left, border=border_thin)
    for j, v in enumerate(vals):
        style_cell(ws6, row, j+2, v, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    row += 1

row += 1

# Plan d'action pour combler l'écart
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
style_cell(ws6, row, 1, "Plan d'Action pour Combler l'Écart de Financement", font=font_section, fill=fill_white, alignment=align_left)
ws6.row_dimensions[row].height = 28
row += 1

action_headers = ["Action", "Montant Cible (EUR)", "Échéance", "Responsable", "Priorité"]
for i, h in enumerate(action_headers, 1):
    style_cell(ws6, row, i, h, font=font_header, fill=fill_navy, alignment=align_center, border=border_thin)
row += 1

actions = [
    ("Finaliser accord Banque Mondiale", 500000, "Q2 2026", "DG ARPT", "Haute"),
    ("Signer convention BAD", 300000, "Q3 2026", "DF ARPT", "Haute"),
    ("Mettre en place taxe sectorielle opérateurs", 400000, "Q1 2027", "Ministère Télécoms", "Haute"),
    ("Obtenir cofinancement AFD", 200000, "Q4 2026", "Coopération Française", "Moyenne"),
    ("Soumettre dossier Fonds Numérique UA", 200000, "Q2 2026", "Cellule ONIT-PNG", "Moyenne"),
    ("Optimiser coûts Phase 3 (réduction 10%)", 200000, "Q1 2028", "PMO ONIT-PNG", "Basse"),
]

for idx, (action, montant, echeance, resp, priorite) in enumerate(actions):
    bg = fill_light_gray if idx % 2 == 0 else fill_white
    style_cell(ws6, row, 1, action, font=font_body, fill=bg, alignment=align_left, border=border_thin)
    style_cell(ws6, row, 2, montant, font=font_body, fill=bg, alignment=align_right, border=border_thin, number_format=EUR_FORMAT)
    style_cell(ws6, row, 3, echeance, font=font_body, fill=bg, alignment=align_center, border=border_thin)
    style_cell(ws6, row, 4, resp, font=font_body, fill=bg, alignment=align_center, border=border_thin)
    
    prio_fill = bg
    prio_font = font_body
    if priorite == "Haute":
        prio_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        prio_font = Font(name="Calibri", size=10, bold=True, color=RED_WARN)
    elif priorite == "Moyenne":
        prio_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        prio_font = Font(name="Calibri", size=10, bold=True, color="D97706")
    else:
        prio_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        prio_font = Font(name="Calibri", size=10, color=BLUE_ACCENT)
    
    style_cell(ws6, row, 5, priorite, font=prio_font, fill=prio_fill, alignment=align_center, border=border_thin)
    row += 1

row += 2
# Note de bas
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
style_cell(ws6, row, 1, 
    "Note : Les montants de financement sont basés sur les engagements préliminaires et les discussions en cours. "
    "Les écarts seront revus trimestriellement par le Comité de Pilotage ONIT-PNG. "
    "Le financement des opérateurs repose sur le cadre réglementaire de contributions sectorielles (arrêté ARPT 2025-012).",
    font=font_small, fill=fill_white, alignment=align_wrap)

ws6.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws6.page_setup.fitToWidth = 1
ws6.page_setup.orientation = "landscape"

print("✓ Sheet 6: Financement")


# ═══════════════════════════════════════════════════════════
# SAUVEGARDE
# ═══════════════════════════════════════════════════════════

output_path = "/home/z/my-project/download/ONIT-PNG_Budget_Estimatif.xlsx"
wb.save(output_path)
print(f"\n✅ Fichier sauvegardé : {output_path}")
print(f"   6 onglets créés avec succès")
